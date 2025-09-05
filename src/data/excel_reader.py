"""
Excel文件读取器

负责读取Excel文件并解析数据
"""

import pandas as pd
import openpyxl
import os
from pathlib import Path

class ExcelReader:
    """
    Excel文件读取器类
    """
    
    def __init__(self, file_path: str):
        """
        初始化Excel读取器
        
        Args:
            file_path: Excel文件路径
        """
        self.file_path = Path(file_path)
        self.data = None
        
    def read_data(self):
        """
        读取Excel数据
        
        Returns:
            解析后的数据
        """
        try:
            if not self.file_path.exists():
                raise FileNotFoundError(f"文件不存在: {self.file_path}")
            
            # 读取Excel文件
            if self.file_path.suffix.lower() == '.xlsx':
                self.data = pd.read_excel(self.file_path, engine='openpyxl')
            elif self.file_path.suffix.lower() == '.xls':
                self.data = pd.read_excel(self.file_path, engine='xlrd')
            else:
                raise ValueError("不支持的文件格式，请使用 .xlsx 或 .xls 文件")
            
            return self.data
        except Exception as e:
            raise Exception(f"读取Excel文件失败: {str(e)}")
    
    def read_data_by_cell_address(self):
        """
        按单元格地址格式读取Excel数据 (A1, B2, C3...)
        
        Returns:
            dict: 单元格地址 -> 值的字典
        """
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(self.file_path)
            worksheet = workbook.active  # 使用第一个工作表
            
            cell_data = {}
            
            # 遍历所有有值的单元格
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_address = cell.coordinate  # 如: A1, B2, C3...
                        cell_data[cell_address] = cell.value
            
            workbook.close()
            return cell_data
            
        except Exception as e:
            raise Exception(f"按单元格地址读取Excel文件失败: {str(e)}")
    
    def get_file_info(self):
        """
        获取文件信息
        
        Returns:
            文件信息字典
        """
        if not self.file_path.exists():
            return None
        
        stat = self.file_path.stat()
        return {
            'name': self.file_path.name,
            'size': stat.st_size,
            'size_mb': round(stat.st_size / (1024*1024), 2),
            'rows': len(self.data) if self.data is not None else 0
        }
    
    def get_columns(self):
        """
        获取Excel文件的列名
        
        Returns:
            列名列表
        """
        if self.data is not None:
            return list(self.data.columns)
        return []
    
    def extract_box_label_data(self, sheet_name=None):
        """
        提取盒标相关的特定单元格数据
        
        Args:
            sheet_name: 工作表名称，如果为None则使用第一个工作表
        
        Returns:
            dict: 包含A4, B4, B11, F4位置数据的字典，并尝试计算结束号
        """
        try:
            if not self.file_path.exists():
                raise FileNotFoundError(f"文件不存在: {self.file_path}")
            
            # 使用openpyxl直接读取特定单元格
            workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            if sheet_name:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # 提取指定单元格数据
            extracted_data = {
                'A4': worksheet['A4'].value,   # 客户名称编码
                'B4': self._find_label_name_content(worksheet),  # 搜索"标签名称:"右边的内容
                'B11': worksheet['B11'].value, # 开始号
                'F4': worksheet['F4'].value,   # 总张数
            }
            
            # 清理数据 - 移除None值并转换为适当格式
            for key, value in extracted_data.items():
                if value is None:
                    extracted_data[key] = ""
                else:
                    extracted_data[key] = str(value).strip()
            
            # 处理总张数，确保是数字
            try:
                if extracted_data['F4']:
                    extracted_data['F4'] = int(float(extracted_data['F4']))
                else:
                    extracted_data['F4'] = 0
            except (ValueError, TypeError):
                extracted_data['F4'] = 0
            
            # 常规模版1：简单的+1递增，不需要复杂分析
            
            workbook.close()
            return extracted_data
            
        except Exception as e:
            raise Exception(f"提取盒标数据失败: {str(e)}")
    
    def _find_label_name_content(self, worksheet):
        """
        在工作表中搜索"标签名称:"关键字，并返回其右边单元格的内容
        
        Args:
            worksheet: openpyxl工作表对象
            
        Returns:
            str: 标签名称内容，如果未找到则返回空字符串
        """
        try:
            # 搜索范围：前20行，前10列（A到J）
            for row in range(1, 21):
                for col in range(1, 11):  # A=1, B=2, ..., J=10
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        # 检查是否包含"标签名称"关键字（不区分是否有冒号）
                        if "标签名称" in cell_value:
                            print(f"找到标签名称关键字: 位置({row},{col}) 内容='{cell_value}'")
                            # 尝试获取右边单元格的内容
                            right_cell = worksheet.cell(row=row, column=col+1).value
                            if right_cell:
                                print(f"找到标签名称右边内容: 位置({row},{col+1}) 内容='{right_cell}'")
                                return str(right_cell).strip()
                            
                            # 如果右边单元格为空，尝试同一个单元格中":"后面的内容
                            if ":" in cell_value:
                                parts = cell_value.split(":", 1)
                                if len(parts) > 1 and parts[1].strip():
                                    content = parts[1].strip()
                                    print(f"从同单元格提取标签名称: '{content}'")
                                    return content
            
            # 如果没找到"标签名称:"，使用B4单元格作为备用方案
            b4_value = worksheet['B4'].value
            if b4_value:
                print(f"未找到'标签名称:'关键字，使用B4单元格内容: '{b4_value}'")
                return str(b4_value).strip()
            
            print("未找到标签名称，使用默认值")
            return "默认主题"
            
        except Exception as e:
            print(f"搜索标签名称失败: {e}")
            # 备用方案：使用B4单元格
            try:
                b4_value = worksheet['B4'].value
                return str(b4_value).strip() if b4_value else "默认主题"
            except:
                return "默认主题"
    
    def get_worksheet_names(self):
        """
        获取Excel文件中所有工作表名称
        
        Returns:
            list: 工作表名称列表
        """
        try:
            if not self.file_path.exists():
                return []
            
            workbook = openpyxl.load_workbook(self.file_path)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
            
        except Exception as e:
            print(f"获取工作表名称失败: {e}")
            return []